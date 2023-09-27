#Scrape imports
import time
import logging
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

#Process imports
import pandas as pd
import re

#Write imports
import os
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter 
from bs4 import BeautifulSoup

headers = {'user-agent': "alanuthuppan@yahoo.com"}

logging.basicConfig(filename='scraping_errors.log', level=logging.ERROR)

def scrape(cik):
    #Final list of soups
    result = []

    #Initialize driver
    base_url = "https://www.sec.gov/"
    landing_url = "edgar/browse/?CIK="
    driver = webdriver.Chrome()
    wait = WebDriverWait(driver, 10)

    try:
        # Navigate to the URL and extract the name of the company
        url = base_url + landing_url + str(cik)
        driver.get(url)
        title = driver.find_element(By.ID, "name")
        name = title.text

        # Locate the expand button and open it
        h5_tags = driver.find_elements(By.TAG_NAME, "h5")
        for h5_tag in h5_tags:
            if h5_tag.text == "[+] 10-K (annual reports) and 10-Q (quarterly reports)":
                h5_tag.click()
                time.sleep(1)
                break

        # Locate the View All button and click it
        xpath = '//button[text()="View all 10-Ks and 10-Qs"]'
        button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        button.click()
        time.sleep(1)

        # Fetch the new page content and loop through the table to get all of the document urls
        document_urls = []
        table = driver.find_element(By.ID, 'filingsTable')
        rows = table.find_elements(By.TAG_NAME, 'tr')
        for row in rows[1:]:
            columns = row.find_elements(By.TAG_NAME, 'td')
            if len(columns) > 1:
                link_element = columns[1].find_element(By.TAG_NAME, 'a')
                if link_element:
                    document_urls.append(link_element.get_attribute('href'))

        # Separate IXBRL documents and HTML documents
        ix_docs = []
        html_docs = []
        for url in document_urls:
            if url.__contains__('ix?'):
                ix_docs.append(url)
            else:
                html_docs.append(url)

        # For IXBRL documents, open them as HTML documents and then add soup contents to list
        for url in ix_docs:
            driver.get(url)
            # Locate and click the "Menu" dropdown
            menu_dropdown = wait.until(EC.element_to_be_clickable((By.ID, "menu-dropdown-link")))
            time.sleep(2)
            menu_dropdown.click()
            

            # Now, within the opened dropdown, locate the "Open as HTML" link and click it
            open_as_html_link = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Open as HTML')]")))
            open_as_html_link.click()
            time.sleep(1)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            result.append((soup, name))

        # For HTML documents, simply add soup contents to list
        for url in html_docs:
            driver.get(url)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            result.append((soup, name))

    except Exception as e:
        logging.error(f"Error scraping CIK {cik}: {e}")
    
    finally:
        driver.close()

    return result


def grab(soup, name):
    # List to store the result (tuples of table and date)
    result = []

    # Search for div containing "Consolidated Schedule of Investments"
    for table in soup.find_all('table'): 
        if ("consolidated schedule of investment" in table.get_text().lower()) and (name.lower() in table.get_text().lower()):
            
            # Find the div/span containing the phrase within the table
            element_containing_phrase = table.find(lambda tag: tag.name in ["div", "span"] and "consolidated schedule of investment" in (tag.get_text() or '').lower())

            # If found and it's a div, get the next div sibling
            if element_containing_phrase and element_containing_phrase.name == 'div':
                next_div = element_containing_phrase.find_next_sibling('div')
                if next_div:
                    date_text = next_div.get_text()
                    result.append((table, date_text.strip()))
            
            # If found and it's a span, split on br tags to get list of strings
            elif element_containing_phrase and element_containing_phrase.name == 'span':
                texts = element_containing_phrase.stripped_strings
                texts_list = list(texts)
                date_text = texts_list[texts_list.index("Consolidated Schedule of Investments") + 1] if "Consolidated Schedule of Investments" in texts_list else None
                if date_text:
                    result.append((table, date_text.strip()))

    return result


def convert_value(value):
    # Convert values with only numbers to integer
    if re.match(r'^[\d,]+$', value):
        return int(value.replace(',', ''))
    
    # Convert percentage formatted values to float (0.xx format)
    elif '%' in value:
        try:
            return round(float(value.replace('%', '')) / 100, 4)
        except ValueError:
            pass

    # Convert float in text form to actual float
    try:
        return round(float(value), 4)
    except ValueError:
        pass

    # If none of the above, return the value as is
    return value


def table_to_dataframe(table, date):
    data = []
    headers = []
    
    rows = table.find_all('tr')[3:]  # Skip the first three rows

    try:
        for header in rows[0].find_all('td'):
            if header.find('span') is None:
                continue
            headers.append(header.get_text().strip())
        data.append(headers)

        for row in rows[1:]:
            # Check if the row contains a cell with a top border
            skip_row = any("border-top" in cell.get("style", "") for cell in row.find_all("td"))
            if skip_row:
                continue
            
            row_data = []
            for cell in row.find_all('td'):
                # Skip cells containing only '$' or '%'
                if cell.get_text().strip() in ["$", "%"]:
                    continue

                # Check if the cell contains a span element, if not skip
                if cell.find('span') is None:
                    continue

                # Skip cell if it contains capitalized, three-letter text
                if len(cell.get_text()) == 3 and cell.get_text().isupper():
                    continue
                
                cell_value = cell.get_text().strip()
                converted_value = convert_value(cell_value)
                row_data.append(converted_value)

            # Combine third and fourth cells and shift left
            if len(row_data) > 4:
                row_data[2] = str(row_data[2]) + '         ' + str(row_data[3])
                del row_data[3]
            
            # Append this row data to the overall data
            if row_data:
                data.append(row_data)
        
        # Convert to DataFrame
        df = pd.DataFrame(data[1:], columns=headers)  # Exclude the first row (headers) from data
    except Exception as e:
        logging.error(f"A DataFrame creation error occurred: {e}")
        df = pd.DataFrame()
        
    return (df, date)



def write_to_excel(tables_by_date):
    filename = "cleaned_soi_tables.xlsx"
    
    # Check if the file exists, if not create a new Excel file
    if not os.path.exists(filename):
        book = Workbook()
        book.save(filename)
    
    # Open the existing Excel file
    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    
    for date, dfs in tables_by_date.items():
        for df in dfs:
            if date in book.sheetnames:
                # If sheet exists, read the existing data and find the next available row (startrow)
                sheet = book[date]
                data = sheet.values
                
                # Check if the sheet is empty
                try:
                    columns = next(data)[0:]
                    existing_df = pd.DataFrame(data, columns=columns)
                    startrow = existing_df.shape[0] + 1
                except StopIteration:
                    # If the sheet is empty, set startrow to 0
                    startrow = 0
                
                # Check if the sheet is empty to determine whether to write the header
                write_header = startrow == 0
                
                # Write the new DataFrame to the existing sheet starting from the next available row
                df.to_excel(writer, sheet_name=date, startrow=startrow, index=False, header=write_header)
            else:
                # If sheet does not exist, create a new sheet with the given date and write DataFrame to it
                df.to_excel(writer, sheet_name=date, index=False)
    
    # Check if the default sheet "Sheet" exists and remove it
    if 'Sheet' in book.sheetnames:
        book.remove(book['Sheet'])

    # Save the changes to the Excel file
    writer.save()
    
    # Reopen the workbook to apply column width, font, and cell alignment adjustments
    book = load_workbook(filename)
    for sheetname in book.sheetnames:
        sheet = book[sheetname]
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            for cell in column:
                cell.alignment = Alignment(horizontal='center')
                if cell.row == 1:
                    cell.font = Font(bold=True)
    
    # Save the final adjustments to the Excel file
    book.save(filename)



# # Scrape the pages from the CIK 
# cik = 1803498
# filing_soups = scrape(cik)

soup1 = BeautifulSoup(requests.get("https://www.sec.gov/Archives/edgar/data/1803498/000180349823000017/bcred-20230630.htm", headers=headers).content, "html.parser")
soup2 =  BeautifulSoup(requests.get("https://www.sec.gov/Archives/edgar/data/1803498/000180349823000012/bcred-20230331.htm", headers=headers).content, "html.parser")
filing_soups = [(soup1, "Blackstone Private Credit Fund"), (soup2, "Blackstone Private Credit Fund")]

# Extract all the relevant tables and dates from each filing

any_errors = False
tables_by_date = {}
for soup, name in filing_soups:
    df_tables = []
    try:
        result = grab(soup, name)
        if result:
            for table, date in result:
                df, _ = table_to_dataframe(table, date)

                if date in tables_by_date:
                    tables_by_date[date].append(df)
                else:
                    tables_by_date[date] = [df]
    except Exception as e:
        print(f"A soup error occurred: {e}")
        any_errors = True

if any_errors:
    print("Some errors occurred while processing the tables. Please check the Excel file for partial results.")
else:
    write_to_excel(tables_by_date)
    print("All tables were processed successfully.") 