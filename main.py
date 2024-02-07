import pandas as pd
import requests
from bs4 import BeautifulSoup
import os
import re
import traceback
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

headers = {'user-agent': "alanuthuppan@yahoo.com"}

def read_filing_links(excel_path):
    df = pd.read_excel(excel_path, sheet_name='Sheet1')
    return df['Filings URL'].tolist()

def normalize_date(date_str):
    try:
        return datetime.strptime(date_str, '%B %d, %Y').strftime('%B %d, %Y')
    except ValueError:
        return date_str

def read_reporting_dates(excel_path):
    dates = []
    df = pd.read_excel(excel_path, sheet_name='Sheet1')
    initial = df['Reporting date'].tolist()
    for date in initial:
        dates.append(str(date.to_pydatetime().strftime('%B %d, %Y').strip()))
    
    return dates

def get_soup_content(url):
    response = requests.get(url, headers=headers)
    print(response.status_code)
    return BeautifulSoup(response.content, 'html.parser')



#EXTRACT

def is_header_div(div, phrase):
    return phrase in div.text.lower() and div.find('table') is not None

def is_data_table_div(div):
    return div.find('table') is not None and len(div.find_all('tr')) > 3

def find_relevant_divs(soup, company_name, phrase, date):
    relevant_divs = []
    all_divs = soup.find_all('div')

    for div in all_divs:
        div_date = get_date_from_div(div, phrase)
        if div_date:
            if phrase in div.text.lower() and company_name in div.text and normalize_date(div_date) == date:
                relevant_divs.append(div)
    print(f"Found {len(relevant_divs)} relevant divs containing '{phrase}' and '{company_name}' and '{date}'")
    return relevant_divs

def get_date_from_div(div, phrase):
    date_pattern = r'\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}\b'

    # Helper function to search for a date pattern in a given text
    def search_date(text):
        return re.search(date_pattern, text)

    # Helper function to process font tags
    def process_font_tags(font_tags):
        for font_tag in font_tags:
            # Split the text in the font tag by <br/> tags
            font_contents = font_tag.decode_contents().split('<br/>')
            for content in font_contents:
                if (match := search_date(content)):
                    return match.group()
        return None
    
    def process_div_tags(div_tags):
        for div_tag in div_tags:
            # Search within the text of the div tag
            if (match := search_date(div_tag.get_text())):
                return match.group()

            # Process any nested div tags
            nested_divs = div_tag.find_all('div', recursive=False)
            if nested_divs:
                date = process_div_tags(nested_divs)
                if date:
                    return date
        return None

    # Check for the phrase in div and its children
    if phrase in div.text.lower():
        # Search in immediate children of the div
        for child in div.children:
            if hasattr(child, 'text') and phrase in child.text.lower():
                date_candidate = child.find_next_sibling()
                if date_candidate and hasattr(date_candidate, 'text'):
                    if (match := search_date(date_candidate.get_text())):
                        return match.group()

        # Search within div/span/font tags
        element_containing_phrase = div.find(lambda tag: tag.name in ["div", "span", "font"] and phrase in (tag.get_text() or '').lower())
        if element_containing_phrase:
            if element_containing_phrase.name == 'div':
                next_div = element_containing_phrase.find_next_sibling('div')
                if next_div and (match := search_date(next_div.get_text())):
                    return match.group()
            else:  # For span or font tags
                texts = element_containing_phrase.stripped_strings
                for text in texts:
                    if (match := search_date(text)):
                        return match.group()

        # Search within table tags and font tags inside tables
        tables = div.find_all('table')
        for table in tables:
            if (match := search_date(table.get_text())):
                return match.group()
            # Process font tags within tables
            font_tags_in_table = table.find_all('font')
            if (date_text := process_font_tags(font_tags_in_table)):
                return date_text

        # Process font tags outside tables
        font_tags_outside_table = div.find_all('font', recursive=False)
        if (date_text := process_font_tags(font_tags_outside_table)):
            return date_text
        
        return process_div_tags([div])

    return None

def extract_tables(soup, company_name, phrase, date):
    relevant_divs = find_relevant_divs(soup, company_name, phrase, date)
    dataframes = []
    extracted_date = None

    for i, div in enumerate(relevant_divs):
        if is_header_div(div, phrase) and is_data_table_div(div):
            print("Found combined")
            table = div.find('table')
            process_table(table, dataframes)
            extracted_date = None
            continue

        if is_header_div(div, phrase):
            print("Header div found, getting table that immediately follows")
            extracted_date = get_date_from_div(div, phrase)
            if extracted_date and normalize_date(extracted_date) == normalize_date(date):
                next_div = div.parent.find_next_sibling()
                if next_div:
                    table = next_div.find('table')
                    if table and "asset" in table.text.lower():
                        process_table(table, dataframes)
                        extracted_date = None
                    elif not next_div.get_text(strip=True):
                        table = next_div.find_next_sibling().find('table')
                        if table and "asset" in table.text.lower():
                            process_table(table, dataframes)
                            extracted_date = None
                    else:
                        print("Following table is not proper data table\n")
                else:
                    print("There is no next div\n")

 
            else:
                print("Date does not match\n")


    return dataframes

#EXTRACT



#PROCESS/CLEAN

def is_subtotal_row(row):
    if row.find('td').get_text(strip=True) == "":
        return True
    else:
        return False

def find_header_row(table):
    rows = table.find_all('tr')
    if not rows:
        print("No header row found")
        print(table)
        return None
    
    for index, row in enumerate(rows):
        cells = row.find_all('td')
        if cells and 'assets' in cells[-1].get_text().lower():
            return (index, row)

    return None

def extract_headers(header_row):
    headers = []
    for header_cell in header_row.find_all('td'):
        # Extract text from each header cell
        header_text = header_cell.get_text(strip=True)
        
        # If the header cell is empty, skip it
        if not header_text:
            continue
        
        # Append the cleaned header text to the headers list
        headers.append(header_text)
    
    return headers

def extract_row_data(row):
    first_cell_text = row.find('td').get_text(strip=True)
    is_bold = 'font-weight:700' in str(row)

    # Determine row type
    if is_subtotal_row(row):
        return None, 'subtotal'
    elif is_bold and 'total' not in first_cell_text.lower():
        return first_cell_text, 'subheader'
    elif is_bold and 'total' in first_cell_text.lower():
        return row, 'total'
    else:
        return row, 'standard'

def process_standard_row(row, headers, current_investment_type, current_industry):
    row_data = [''] * len(headers)
    td_elements = row.find_all('td')
    columns_filled = [False] * len(headers)

    # Handle the Investment Name (first cell)
    investment_name = td_elements[0].get_text(strip=True)
    row_data[0] = investment_name
    columns_filled[0] = True

    # Iterate from right to left, skipping the first cell
    for i in range(len(td_elements) - 1, 0, -1):
        td = td_elements[i]
        cell_text = td.get_text(strip=True)

        if not cell_text or cell_text == '%':
            continue

        # Check for currency/country codes and prepend to the last numeric value
        if re.match(r'^[A-Z]{3}$', cell_text) or cell_text in ['$', '£', 'EUR']:
            # Find the last filled numeric value
            last_numeric_index = next((j for j, val in enumerate(row_data) if val.replace(',', '').replace('.', '', 1).isdigit() and columns_filled[j]), None)
            if last_numeric_index is not None:
                row_data[last_numeric_index] = cell_text + ' ' + row_data[last_numeric_index]
            continue

        # Assign cell text to appropriate column based on the pattern
        if cell_text.replace('.', '', 1).replace(',', '').isdigit():
            # Fill in numeric columns in the order: Percentage of Net Assets, Fair Value, Cost, Par Amount/Units
            for header in reversed(['amount', 'cost', 'fair', 'assets']):
                index = headers.index(next(filter(lambda h: header in h.lower(), headers), None))
                if not columns_filled[index]:
                    row_data[index] = cell_text
                    columns_filled[index] = True
                    break

        # Interest Rate and Spread
        elif '%' in cell_text:
            if 'interest' in ''.join(headers).lower() and not columns_filled[headers.index(next(filter(lambda h: 'interest' in h.lower(), headers), None))]:
                # If Interest Rate not filled, fill it
                index = headers.index(next(filter(lambda h: 'interest' in h.lower(), headers), None))
                row_data[index] = cell_text
                columns_filled[index] = True
            elif 'reference' in ''.join(headers).lower() and not columns_filled[headers.index(next(filter(lambda h: 'reference' in h.lower(), headers), None))]:
                # Fill Spread
                index = headers.index(next(filter(lambda h: 'reference' in h.lower(), headers), None))
                row_data[index] = cell_text
                columns_filled[index] = True

        # Reference Rate
        elif '+' in cell_text and 'reference' in ''.join(headers).lower():
            index = headers.index(next(filter(lambda h: 'reference' in h.lower(), headers), None))
            if columns_filled[index]:
                # Prepend Reference Rate to existing Spread value
                row_data[index] = cell_text + ' ' + row_data[index]

        # Maturity Date
        elif re.match(r'\d{1,2}/\d{1,2}/\d{4}', cell_text):
            index = headers.index(next(filter(lambda h: 'maturity' in h.lower(), headers), None))
            row_data[index] = cell_text
            columns_filled[index] = True

        # Footnotes
        elif '(' in cell_text and 'footnotes' in ''.join(headers).lower():
            index = headers.index(next(filter(lambda h: 'footnote' in h.lower(), headers), None))
            row_data[index] = cell_text
            columns_filled[index] = True

    row_data[1] = current_investment_type
    row_data[2] = current_industry

    if len(row_data) != len(headers):
        print("Mismatch in headers and row data lengths")
        print("Headers:", headers)
        print("Row Data:", row_data)
        raise ValueError("Row data does not align with headers")

    return row_data


def append_row_to_df(data, row_data):
    # Append the processed row data to the main data list
    data.append(row_data)
    return data

def clean(df):
    df = df.astype(str)
    df.replace({'nan': '', 'None': '', 'NaN': ''}, inplace=True)

    return df

def process_table(table, dfs):
    header_row_info = find_header_row(table)
    if not header_row_info:
        print("No header row found in table.")
        
    header_row_idx, header_row = header_row_info
    headers = extract_headers(header_row)
    headers.insert(1, 'Industry')
    headers.insert(1, 'Investment Type')    

    data = []
    current_investment_type = ""
    current_industry = ""

    rows = table.find_all('tr')[header_row_idx + 1:]
    for row in rows:
        row_content, row_type = extract_row_data(row)
        
        if row_type == 'subtotal':
            continue
        elif row_type == 'subheader':
            if any(word in row_content.lower() for word in ['controlled', 'affiliated', 'debt', 'lien', 'equity', 'structured']) and 'total' not in row_content.lower():
                current_investment_type = row_content.replace("(continued)", "").strip()
            else:
                current_industry = row_content.replace("(continued)", "").strip()
            continue
        elif row_type == 'total':
            current_investment_type = ""
            current_industry = ""
            row_data = process_standard_row(row_content, headers, current_investment_type, current_industry)
        elif row_type == 'standard':
            row_data = process_standard_row(row_content, headers, current_investment_type, current_industry)

        data = append_row_to_df(data, row_data)

    df = pd.DataFrame(data, columns=headers)
    df = clean(df)
    dfs.append(df)
    print("Table processed succesfully\n")


#PROCESS/CLEAN



#WRITE

def append_df_to_excel(filename, df, sheet_name, startrow=None, **to_excel_kwargs):
    if df.empty:
        print(f"DataFrame is empty for sheet: {sheet_name}. Skipping.")
        return
    
    # Create a new Excel file if it doesn't exist
    if not os.path.isfile(filename):
        df.to_excel(filename, sheet_name=sheet_name, startrow=startrow if startrow is not None else 0, index=False, header=True, **to_excel_kwargs)
    else:
        # Append to the existing file
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Check existing sheet names
            existing_sheets = writer.book.sheetnames

            # Determine start row for appending
            if sheet_name in existing_sheets:
                startrow = writer.book[sheet_name].max_row
                df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=False, **to_excel_kwargs)
            else:
                # Create a new sheet if it doesn't exist
                df.to_excel(writer, sheet_name, startrow=0, index=False, header=True, **to_excel_kwargs)


def format_workbook(filename):
    workbook = openpyxl.load_workbook(filename)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Freeze the top row
        sheet.freeze_panes = 'A2'

        # Bold 'total' rows and autofit column width
        for row in sheet.iter_rows():
            if 'total' in str(row[0].value).lower():
                for cell in row:
                    cell.font = Font(bold=True)
        
        for col in sheet.iter_cols():
            column_letter = get_column_letter(col[0].column)
            max_length = max(len(str(cell.value)) for cell in col if cell.value) or 10
            sheet.column_dimensions[column_letter].width = max_length

    workbook.save(filename)


#WRITE



#MAIN

def main():
    filings_path = 'all_filings.xlsx'
    urls = read_filing_links(filings_path)
    dates = read_reporting_dates(filings_path)
    output_excel = 'cleaned_soi_tables.xlsx'

    filings = zip(urls, dates)

    for url, date in filings:
        print(f"Processing URL: {url}")
        print(f"Date: {date}")
        try:
            soup = get_soup_content(url)
            dataframes = extract_tables(soup, "Blackstone Private Credit Fund", "consolidated schedule of investment", date)

            if dataframes:
                combined_df = pd.concat(dataframes, ignore_index=True)
                append_df_to_excel(output_excel, combined_df, sheet_name=date)
                print(f"Appended all data for {date} to {output_excel}")

            else:
                print("No data to append for this filing.")

            
        except Exception as e:
            error_traceback = traceback.format_exc()  # Get detailed traceback
            with open('error_log.txt', 'a') as f:
                f.write(f"Error processing {url}:\n{error_traceback}\n\n")
    
    format_workbook(output_excel)
            


if __name__ == "__main__":
    main()